# make_target_rule_table.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.data_preprocessing.project_paths import TABLES_DIR

# ============================================================
# 1. 경로 설정
# ============================================================
OUTPUT_DIR = TABLES_DIR
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_FILE = OUTPUT_DIR / "target_rule_table.xlsx"


# ============================================================
# 2. README 시트 데이터
# ============================================================
summary_rows = [
    {
        "item": "파일명",
        "content": "target_rule_table.xlsx",
    },
    {
        "item": "파일 목적",
        "content": "현재 액상형 전자담배 사용 여부 예측을 위한 타깃 변수 생성 기준을 정리한 문서",
    },
    {
        "item": "원본 타깃 후보 변수",
        "content": "TC_EC_MN",
    },
    {
        "item": "생성 타깃 변수명",
        "content": "current_ecig_use",
    },
    {
        "item": "타깃 정의",
        "content": "최근 30일 동안 액상형 전자담배를 1일 이상 사용했으면 1, 그렇지 않으면 0",
    },
    {
        "item": "y=0 기준",
        "content": "TC_EC_MN=9999 또는 TC_EC_MN=1",
    },
    {
        "item": "y=1 기준",
        "content": "TC_EC_MN=2~7",
    },
    {
        "item": "9999 처리 근거",
        "content": (
            "원시자료 이용지침서의 분기형 문항 비해당 코드 설명과 "
            "TC_EC_LT x TC_EC_MN 교차표 확인 결과를 함께 근거로 사용"
        ),
    },
    {
        "item": "주의사항",
        "content": "본 타깃은 평생 사용 경험 여부가 아니라 현재 사용 여부를 의미한다.",
    },
]

summary_df = pd.DataFrame(summary_rows)


# ============================================================
# 3. 타깃 생성 기준표 데이터
# ============================================================
target_rule_rows = [
    {
        "source_variable": "TC_EC_MN",
        "source_value": "9999",
        "target_variable": "current_ecig_use",
        "target_value": 0,
        "classification": "현재 비사용자",
        "meaning": "분기형 문항의 비해당 코드",
        "decision_reason": (
            "원시자료 이용지침서에서 분기형 적용 문항의 비해당 코드가 9999로 처리될 수 있음을 확인하였다. "
            "또한 TC_EC_LT x TC_EC_MN 교차표에서 TC_EC_LT=1인 응답자에게만 TC_EC_MN=9999가 나타났으므로, "
            "액상형 전자담배 평생 사용 경험이 없어 최근 30일 사용 문항에 해당하지 않는 응답자로 판단하였다."
        ),
        "modeling_decision": "현재 사용자가 아니므로 y=0으로 변환",
    },
    {
        "source_variable": "TC_EC_MN",
        "source_value": "1",
        "target_variable": "current_ecig_use",
        "target_value": 0,
        "classification": "현재 비사용자",
        "meaning": "최근 30일 동안 액상형 전자담배 사용 없음",
        "decision_reason": (
            "TC_EC_MN은 최근 30일 동안 액상형 전자담배를 사용한 날을 묻는 변수이다. "
            "값 1은 최근 30일 동안 사용하지 않았음을 의미하므로, "
            "평생 사용 경험이 있더라도 현재 사용자는 아니다."
        ),
        "modeling_decision": "최근 30일 사용자가 아니므로 y=0으로 변환",
    },
    {
        "source_variable": "TC_EC_MN",
        "source_value": "2~7",
        "target_variable": "current_ecig_use",
        "target_value": 1,
        "classification": "현재 사용자",
        "meaning": "최근 30일 동안 1일 이상 액상형 전자담배 사용",
        "decision_reason": (
            "현재 액상형 전자담배 사용 여부는 최근 30일 동안 1일 이상 사용했는지를 기준으로 정의한다. "
            "TC_EC_MN 값 2~7은 최근 30일 내 사용일수가 1일 이상인 응답 범주에 해당한다."
        ),
        "modeling_decision": "현재 사용자로 보고 y=1로 변환",
    },
]

target_rule_df = pd.DataFrame(target_rule_rows)


# ============================================================
# 4. 보고서 작성용 문장 데이터
# ============================================================
report_text_rows = [
    {
        "section": "타깃 변수 정의 문장",
        "text": (
            "본 연구의 타깃 변수는 현재 액상형 전자담배 사용 여부이다. "
            "원본 변수 TC_EC_MN은 최근 30일 동안 액상형 전자담배를 사용한 날을 묻는 문항이며, "
            "본 연구에서는 TC_EC_MN=2~7을 최근 30일 동안 1일 이상 사용한 현재 사용자(y=1)로, "
            "TC_EC_MN=1을 최근 30일 동안 사용하지 않은 현재 비사용자(y=0)로 분류하였다. "
            "또한 TC_EC_MN=9999는 이용지침서상 분기형 문항의 비해당 코드로 해석할 수 있으며, "
            "TC_EC_LT와의 교차표 확인 결과 평생 액상형 전자담배 사용 경험이 없는 응답자에게만 나타났으므로 "
            "현재 비사용자(y=0)로 처리하였다."
        ),
    },
    {
        "section": "주의 문장",
        "text": (
            "본 연구는 액상형 전자담배의 평생 사용 경험 여부가 아니라 현재 사용 여부를 예측 대상으로 삼았으므로, "
            "평생 사용 경험이 있더라도 최근 30일 동안 사용하지 않은 응답자는 현재 비사용자로 분류하였다."
        ),
    },
]

report_text_df = pd.DataFrame(report_text_rows)


# ============================================================
# 5. 엑셀 저장
# ============================================================
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    summary_df.to_excel(
        writer,
        sheet_name="README",
        index=False,
        startrow=4,
    )

    target_rule_df.to_excel(
        writer,
        sheet_name="target_rule",
        index=False,
        startrow=4,
    )

    report_text_df.to_excel(
        writer,
        sheet_name="report_text",
        index=False,
        startrow=4,
    )


# ============================================================
# 6. 엑셀 서식 설정
# ============================================================
wb = load_workbook(OUTPUT_FILE)

title_fill = PatternFill("solid", fgColor="1F4E78")
subtitle_fill = PatternFill("solid", fgColor="D9EAF7")
header_fill = PatternFill("solid", fgColor="D9EAF7")

white_font = Font(color="FFFFFF", bold=True, size=14)
subtitle_font = Font(bold=True, size=11)
header_font = Font(bold=True)
normal_font = Font(size=10)

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_sheet(ws, title: str, subtitle: str, header_row: int = 5):
    """
    각 시트에 제목, 설명, 헤더 스타일, 본문 스타일, 열 너비를 적용한다.

    주의:
    제목/설명 행을 merge_cells로 병합하면 일부 셀이 MergedCell이 된다.
    따라서 열 너비 조정 시 ws.columns를 직접 순회하지 않고,
    column index 기준으로 접근한다.
    """
    max_col = ws.max_column
    max_row = ws.max_row

    # ------------------------------------------------------------
    # 제목 행
    # ------------------------------------------------------------
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max_col,
    )

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.fill = title_fill
    title_cell.font = white_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ------------------------------------------------------------
    # 설명 행
    # ------------------------------------------------------------
    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=max_col,
    )

    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = subtitle
    subtitle_cell.fill = subtitle_fill
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    ws.row_dimensions[2].height = 35

    # ------------------------------------------------------------
    # 헤더 스타일
    # ------------------------------------------------------------
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = thin_border

    # ------------------------------------------------------------
    # 본문 스타일
    # ------------------------------------------------------------
    for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row):
        for cell in row:
            cell.font = normal_font
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            cell.border = thin_border

    # ------------------------------------------------------------
    # 틀 고정
    # ------------------------------------------------------------
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ------------------------------------------------------------
    # 필터
    # ------------------------------------------------------------
    ws.auto_filter.ref = ws.dimensions

    # ------------------------------------------------------------
    # 열 너비 조정
    # 병합된 1~2행은 제외하고 실제 표 영역 기준으로 계산
    # ------------------------------------------------------------
    for col_idx in range(1, max_col + 1):
        col_letter = ws.cell(row=header_row, column=col_idx).column_letter
        max_length = 0

        for row_idx in range(header_row, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if cell.value is None:
                continue

            value_length = len(str(cell.value))

            if value_length > max_length:
                max_length = value_length

        ws.column_dimensions[col_letter].width = min(
            max(max_length + 2, 12),
            60,
        )


style_sheet(
    wb["README"],
    title="타깃 생성 기준표",
    subtitle="현재 액상형 전자담배 사용 여부 예측을 위한 타깃 변수 생성 기준 요약",
)

style_sheet(
    wb["target_rule"],
    title="TC_EC_MN 기반 타깃 변환 규칙",
    subtitle="TC_EC_MN 원본값을 current_ecig_use 이진 타깃으로 변환하는 기준",
)

style_sheet(
    wb["report_text"],
    title="보고서 작성용 문장",
    subtitle="보고서의 데이터 전처리 또는 타깃 변수 정의 절에 사용할 수 있는 문장",
)

wb.save(OUTPUT_FILE)


# ============================================================
# 7. 실행 결과 출력
# ============================================================
print("저장 완료:", OUTPUT_FILE)

print("\n생성된 시트:")
for sheet_name in wb.sheetnames:
    print("-", sheet_name)

print("\n타깃 생성 기준:")
print(target_rule_df.to_string(index=False))
