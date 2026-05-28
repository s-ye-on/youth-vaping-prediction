# make_final_processed_files.py

from collections import Counter

import pandas as pd
import pyreadstat
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.data_preprocessing.project_paths import PROCESSED_DATA_DIR, TABLES_DIR, sav_files


# ============================================================
# 1. 경로 설정
# ============================================================
DATA_DIR = PROCESSED_DATA_DIR
OUTPUT_TABLE_DIR = TABLES_DIR

DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_TABLE_DIR.mkdir(parents=True, exist_ok=True)

INPUT_CSV = DATA_DIR / "combined_kyrbs_2023_2025.csv"

HUMAN_READABLE_XLSX = DATA_DIR / "combined_kyrbs_2023_2025.xlsx"
MODELING_CSV = DATA_DIR / "modeling_dataset_encoded.csv"
MODELING_XLSX = DATA_DIR / "modeling_dataset_encoded.xlsx"
ENCODING_MAP_XLSX = OUTPUT_TABLE_DIR / "encoding_map.xlsx"

SAV_FILES = sav_files()

TARGET_NAME = "current_ecig_use"


# ============================================================
# 2. 파일 존재 확인
# ============================================================
if not INPUT_CSV.exists():
    raise FileNotFoundError(
        f"통합 CSV 파일을 찾을 수 없습니다: {INPUT_CSV}\n"
        "먼저 build_combined_dataset.py를 실행해서 combined_kyrbs_2023_2025.csv를 생성하세요."
    )

missing_sav_files = [str(path) for path in SAV_FILES.values() if not path.exists()]

if missing_sav_files:
    raise FileNotFoundError(
        "다음 .sav 파일을 찾을 수 없습니다.\n"
        + "\n".join(missing_sav_files)
        + "\n\n원본 .sav에서 변수 라벨을 읽어와야 하므로 sav 파일이 프로젝트 루트에 있어야 합니다."
    )


# ============================================================
# 3. 통합 CSV 읽기
# ============================================================
df = pd.read_csv(INPUT_CSV)

print("통합 CSV 읽기 완료:", INPUT_CSV)
print("데이터 크기:", df.shape)
print("컬럼 목록:")
print(list(df.columns))


# ============================================================
# 4. 원본 SAV에서 변수 라벨 읽기
# ============================================================
label_by_year = {}

for year, path in SAV_FILES.items():
    print(f"{year}년 SAV 메타데이터 읽는 중:", path.name)

    _, meta = pyreadstat.read_sav(
        str(path),
        metadataonly=True,
    )

    label_by_year[year] = meta.column_names_to_labels


def choose_final_label(variable: str) -> str:
    """
    2023~2025년 라벨 중 대표 한글 설명을 선택한다.
    라벨이 연도별로 같으면 그 값을 사용하고,
    다르면 가장 많이 등장한 라벨을 사용한다.
    """
    if variable == TARGET_NAME:
        return "현재 액상형 전자담배 사용 여부"

    if variable == "YEAR":
        return "조사연도"

    labels = []

    for year in sorted(label_by_year.keys()):
        label = label_by_year[year].get(variable)

        if label is not None and str(label).strip() != "":
            labels.append(str(label).strip())

    if not labels:
        return variable

    counter = Counter(labels)
    most_common_label, _ = counter.most_common(1)[0]

    return most_common_label


def get_label_for_year(variable: str, year: int) -> str:
    if variable == TARGET_NAME:
        return "현재 액상형 전자담배 사용 여부"

    if variable == "YEAR":
        return "조사연도"

    return label_by_year.get(year, {}).get(variable, "")


# ============================================================
# 5. feature 설명표 만들기
# ============================================================
feature_description_rows = []

for order, col in enumerate(df.columns, start=1):
    label_2023 = get_label_for_year(col, 2023)
    label_2024 = get_label_for_year(col, 2024)
    label_2025 = get_label_for_year(col, 2025)

    labels = [
        str(label).strip()
        for label in [label_2023, label_2024, label_2025]
        if label is not None and str(label).strip() != ""
    ]

    unique_labels = sorted(set(labels))

    if col == TARGET_NAME:
        role = "target"
    else:
        role = "feature"

    if len(unique_labels) <= 1:
        label_consistency = "same"
    else:
        label_consistency = "different"

    feature_description_rows.append({
        "column_order": order,
        "variable": col,
        "final_korean_label": choose_final_label(col),
        "label_2023": label_2023,
        "label_2024": label_2024,
        "label_2025": label_2025,
        "label_consistency": label_consistency,
        "role": role,
    })

feature_description_df = pd.DataFrame(feature_description_rows)

column_to_label = {
    row["variable"]: row["final_korean_label"]
    for _, row in feature_description_df.iterrows()
}


# ============================================================
# 6. 타깃 규칙표
# ============================================================
target_rule_df = pd.DataFrame([
    {
        "source_variable": "TC_EC_MN",
        "source_value": "9999",
        "target_variable": TARGET_NAME,
        "target_value": 0,
        "meaning": "분기형 문항의 비해당 코드",
        "decision": "현재 비사용자",
    },
    {
        "source_variable": "TC_EC_MN",
        "source_value": "1",
        "target_variable": TARGET_NAME,
        "target_value": 0,
        "meaning": "최근 30일 동안 액상형 전자담배 사용 없음",
        "decision": "현재 비사용자",
    },
    {
        "source_variable": "TC_EC_MN",
        "source_value": "2~7",
        "target_variable": TARGET_NAME,
        "target_value": 1,
        "meaning": "최근 30일 동안 1일 이상 액상형 전자담배 사용",
        "decision": "현재 사용자",
    },
])


# ============================================================
# 7. 타깃 분포표
# ============================================================
target_counts = df[TARGET_NAME].value_counts().sort_index()
target_ratios = df[TARGET_NAME].value_counts(normalize=True).sort_index()

target_distribution_df = pd.DataFrame({
    TARGET_NAME: target_counts.index,
    "count": target_counts.values,
    "ratio": target_ratios.values,
})

year_target_distribution_df = (
    df.groupby(["YEAR", TARGET_NAME])
    .size()
    .reset_index(name="count")
)

year_total = (
    df.groupby("YEAR")
    .size()
    .reset_index(name="year_total")
)

year_target_distribution_df = year_target_distribution_df.merge(
    year_total,
    on="YEAR",
    how="left",
)

year_target_distribution_df["ratio"] = (
    year_target_distribution_df["count"]
    / year_target_distribution_df["year_total"]
)


# ============================================================
# 8. 9999 값 점검표
# ============================================================
missing_code_rows = []

for col in df.columns:
    if pd.api.types.is_numeric_dtype(df[col]):
        count_9999 = int((df[col] == 9999).sum())
    else:
        count_9999 = 0

    if count_9999 > 0:
        missing_code_rows.append({
            "variable": col,
            "korean_label": column_to_label.get(col, col),
            "count_9999": count_9999,
            "ratio_9999": count_9999 / len(df),
            "recommended_handling": (
                "타깃이 아닌 feature의 9999는 무조건 0/1로 바꾸지 말고 "
                "unknown/missing 범주로 처리한다. 모델링용 데이터에서는 -1로 치환한다."
            ),
        })

missing_code_df = pd.DataFrame(missing_code_rows)


# ============================================================
# 9. 모델링용 데이터 생성
# ============================================================
modeling_df = df.copy()

feature_cols = [col for col in modeling_df.columns if col != TARGET_NAME]

# feature의 9999는 -1로 치환
for col in feature_cols:
    if pd.api.types.is_numeric_dtype(modeling_df[col]):
        modeling_df[col] = modeling_df[col].replace(9999, -1)

# 문자열 범주형 변수 숫자 인코딩
encoding_map_rows = []

for col in feature_cols:
    if modeling_df[col].dtype == "object":
        original_series = modeling_df[col].astype("category")
        categories = list(original_series.cat.categories)

        for code, category in enumerate(categories):
            encoding_map_rows.append({
                "variable": col,
                "korean_label": column_to_label.get(col, col),
                "original_value": category,
                "encoded_value": code,
            })

        modeling_df[col] = original_series.cat.codes

# 결측은 -1로 채움
modeling_df = modeling_df.fillna(-1)
modeling_df[TARGET_NAME] = modeling_df[TARGET_NAME].astype(int)

encoding_map_df = pd.DataFrame(encoding_map_rows)


# ============================================================
# 10. 사람이 보기 좋은 전체 데이터 Excel 저장
# ============================================================
with pd.ExcelWriter(HUMAN_READABLE_XLSX, engine="openpyxl") as writer:
    # startrow=1로 저장하면:
    # 1행은 비워두고,
    # 2행에 변수명,
    # 3행부터 데이터가 들어간다.
    # 이후 openpyxl로 1행에 한글 라벨을 넣는다.
    df.to_excel(
        writer,
        sheet_name="combined_data",
        index=False,
        startrow=1,
    )

    feature_description_df.to_excel(
        writer,
        sheet_name="feature_description",
        index=False,
    )

    target_rule_df.to_excel(
        writer,
        sheet_name="target_rule",
        index=False,
    )

    target_distribution_df.to_excel(
        writer,
        sheet_name="target_distribution",
        index=False,
    )

    year_target_distribution_df.to_excel(
        writer,
        sheet_name="year_target_distribution",
        index=False,
    )

    if not missing_code_df.empty:
        missing_code_df.to_excel(
            writer,
            sheet_name="missing_code_9999",
            index=False,
        )


# ============================================================
# 11. Excel combined_data 시트 상단에 한글 라벨 행 추가
# ============================================================
wb = load_workbook(HUMAN_READABLE_XLSX)
ws = wb["combined_data"]

header_fill = PatternFill("solid", fgColor="1F4E78")
variable_fill = PatternFill("solid", fgColor="D9EAF7")
white_font = Font(color="FFFFFF", bold=True)
bold_font = Font(bold=True)

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 1행: 한글 설명
# 2행: 변수명
for col_idx, variable_name in enumerate(df.columns, start=1):
    korean_label = column_to_label.get(variable_name, variable_name)

    label_cell = ws.cell(row=1, column=col_idx)
    label_cell.value = korean_label
    label_cell.fill = header_fill
    label_cell.font = white_font
    label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_cell.border = thin_border

    variable_cell = ws.cell(row=2, column=col_idx)
    variable_cell.fill = variable_fill
    variable_cell.font = bold_font
    variable_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    variable_cell.border = thin_border

# 틀 고정: 1~2행 고정
ws.freeze_panes = "A3"

# 필터는 변수명 행인 2행 기준으로 적용
ws.auto_filter.ref = ws.dimensions

# 열 너비 조정
for col_idx, variable_name in enumerate(df.columns, start=1):
    korean_label = column_to_label.get(variable_name, variable_name)
    max_len = max(len(str(variable_name)), len(str(korean_label)))
    width = min(max(max_len + 4, 12), 35)
    col_letter = ws.cell(row=2, column=col_idx).column_letter
    ws.column_dimensions[col_letter].width = width

# 1~2행 높이 조정
ws.row_dimensions[1].height = 42
ws.row_dimensions[2].height = 28

wb.save(HUMAN_READABLE_XLSX)

print("사람이 보기 좋은 통합 Excel 저장 완료:", HUMAN_READABLE_XLSX)


# ============================================================
# 12. 모델링용 데이터 저장
# ============================================================
modeling_df.to_csv(
    MODELING_CSV,
    index=False,
    encoding="utf-8-sig",
)

with pd.ExcelWriter(MODELING_XLSX, engine="openpyxl") as writer:
    modeling_df.to_excel(
        writer,
        sheet_name="modeling_data",
        index=False,
    )

print("모델링용 CSV 저장 완료:", MODELING_CSV)
print("모델링용 Excel 저장 완료:", MODELING_XLSX)


# ============================================================
# 13. 인코딩 매핑표 저장
# ============================================================
with pd.ExcelWriter(ENCODING_MAP_XLSX, engine="openpyxl") as writer:
    if not encoding_map_df.empty:
        encoding_map_df.to_excel(
            writer,
            sheet_name="encoding_map",
            index=False,
        )

    feature_description_df.to_excel(
        writer,
        sheet_name="feature_description",
        index=False,
    )

    if not missing_code_df.empty:
        missing_code_df.to_excel(
            writer,
            sheet_name="missing_code_9999",
            index=False,
        )

print("인코딩 매핑표 저장 완료:", ENCODING_MAP_XLSX)


# ============================================================
# 14. 결과 출력
# ============================================================
print("\n===== 생성 완료 =====")
print("1. 사람이 보기 좋은 전체 통합 Excel:")
print("   ", HUMAN_READABLE_XLSX)

print("2. 모델링용 CSV:")
print("   ", MODELING_CSV)

print("3. 모델링용 Excel:")
print("   ", MODELING_XLSX)

print("4. 인코딩 매핑표:")
print("   ", ENCODING_MAP_XLSX)

print("\n원본 통합 데이터 크기:", df.shape)
print("모델링용 데이터 크기:", modeling_df.shape)

print("\n타깃 분포:")
print(target_distribution_df.to_string(index=False))

if not missing_code_df.empty:
    print("\n9999 값이 남아 있던 변수:")
    print(missing_code_df.to_string(index=False))
else:
    print("\n9999 값이 남아 있는 변수는 없습니다.")
