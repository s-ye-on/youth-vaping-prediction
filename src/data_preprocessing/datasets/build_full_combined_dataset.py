# build_full_combined_dataset.py

# 전체 원자료 병합, 전체 feature 확인
from collections import Counter

import pandas as pd
import pyreadstat
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.data_preprocessing.project_paths import PROCESSED_DATA_DIR, PROJECT_ROOT, TABLES_DIR, sav_files


# ============================================================
# 1. 경로 설정
# ============================================================
SAV_FILES = sav_files()

DATA_DIR = PROCESSED_DATA_DIR
TABLE_DIR = TABLES_DIR

DATA_DIR.mkdir(parents=True, exist_ok=True)
TABLE_DIR.mkdir(parents=True, exist_ok=True)

FULL_CSV_FILE = DATA_DIR / "full_combined_kyrbs_2023_2025.csv"
FULL_XLSX_FILE = DATA_DIR / "full_combined_kyrbs_2023_2025.xlsx"
FEATURE_CATALOG_FILE = TABLE_DIR / "full_feature_catalog_2023_2025.xlsx"

TARGET_SOURCE = "TC_EC_MN"
TARGET_NAME = "current_ecig_use"


# ============================================================
# 2. 파일 존재 확인
# ============================================================
missing_files = [str(path) for path in SAV_FILES.values() if not path.exists()]

if missing_files:
    raise FileNotFoundError(
        "다음 .sav 파일을 찾을 수 없습니다.\n"
        + "\n".join(missing_files)
        + "\n\n확인 사항:\n"
        + "1. kyrbs2023.sav, kyrbs2024.sav, kyrbs2025.sav가 프로젝트 루트에 있는지 확인하세요.\n"
        + "2. 파일명이 정확히 일치하는지 확인하세요."
    )


# ============================================================
# 3. 타깃 생성 함수
# ============================================================
def make_current_ecig_use(value):
    """
    현재 액상형 전자담배 사용 여부 생성 규칙

    TC_EC_MN = 9999 -> 비해당, 현재 비사용자, 0
    TC_EC_MN = 1    -> 최근 30일 동안 사용 없음, 0
    TC_EC_MN = 2~7  -> 최근 30일 동안 1일 이상 사용, 1
    """
    if pd.isna(value):
        return pd.NA

    value = int(value)

    if value in [9999, 1]:
        return 0

    if value in [2, 3, 4, 5, 6, 7]:
        return 1

    return pd.NA


# ============================================================
# 4. SAV 메타데이터 읽기
# ============================================================
metadata_by_year = {}
columns_by_year = {}
labels_by_year = {}
value_labels_by_year = {}

for year, path in SAV_FILES.items():
    print(f"{year}년 메타데이터 읽는 중: {path.name}")

    _, meta = pyreadstat.read_sav(
        str(path),
        metadataonly=True,
    )

    metadata_by_year[year] = meta
    columns_by_year[year] = list(meta.column_names)
    labels_by_year[year] = dict(meta.column_names_to_labels)
    value_labels_by_year[year] = dict(meta.variable_value_labels)


# ============================================================
# 5. 전체 feature 목록 생성
# ============================================================
all_columns = sorted(
    set().union(*[set(cols) for cols in columns_by_year.values()])
)

print("\n연도별 원본 변수 개수:")
for year in sorted(columns_by_year):
    print(f"{year}: {len(columns_by_year[year])}개")

print("\n3개년 합집합 변수 개수:", len(all_columns))

common_columns = sorted(
    set(columns_by_year[2023])
    & set(columns_by_year[2024])
    & set(columns_by_year[2025])
)

print("3개년 공통 변수 개수:", len(common_columns))


def choose_label(variable):
    labels = []

    for year in sorted(labels_by_year):
        label = labels_by_year[year].get(variable)

        if label is not None and str(label).strip() != "":
            labels.append(str(label).strip())

    if not labels:
        return ""

    return Counter(labels).most_common(1)[0][0]


def label_consistency(variable):
    labels = []

    for year in sorted(labels_by_year):
        label = labels_by_year[year].get(variable)

        if label is not None and str(label).strip() != "":
            labels.append(str(label).strip())

    unique_labels = sorted(set(labels))

    if len(unique_labels) <= 1:
        return "same"

    return "different"


def infer_domain(variable):
    if variable in [
        "YEAR", "CITY", "CTYPE", "CTYPE_SD", "MH", "SCHOOL", "STYPE",
        "STRATA", "STRATA_NM", "CLUSTER", "FPC", "GROUP", "W", "OBS", "mod_d"
    ]:
        return "common_or_design"

    prefix_map = {
        "TC": "흡연",
        "AC": "음주",
        "WC": "비만/체중조절",
        "PA": "신체활동",
        "F": "식생활",
        "I": "손상예방",
        "V": "폭력",
        "S": "성행태",
        "M": "정신건강",
        "O": "구강건강",
        "HW": "개인위생",
        "AS": "천식",
        "RH": "알레르기비염",
        "ECZ": "아토피피부염",
        "INT": "인터넷중독",
        "DR": "약물",
        "E": "건강형평성/가정환경",
    }

    for prefix, domain in prefix_map.items():
        if variable.startswith(prefix + "_"):
            return domain

    return "unknown"


def infer_initial_decision(variable):
    """
    전체 feature 탐색용 1차 분류.
    최종 feature 결정이 아니라 후보 검토를 쉽게 하기 위한 임시 태그다.
    """
    if variable == TARGET_NAME:
        return "target_created"

    if variable == TARGET_SOURCE:
        return "target_source_keep_for_audit_not_feature"

    if variable.startswith("TC_EC_"):
        return "exclude_leakage_candidate"

    if variable.startswith("TC_"):
        return "review_tobacco_related_possible_leakage"

    if variable in [
        "STRATA", "STRATA_NM", "CLUSTER", "FPC", "GROUP", "W", "OBS", "mod_d"
    ]:
        return "exclude_design_or_admin"

    if variable in ["YEAR", "CITY", "CTYPE", "MH", "SCHOOL", "STYPE"]:
        return "candidate_basic"

    if variable.startswith("E_"):
        return "candidate_family_or_equity_review"

    return "candidate_review"


feature_rows = []

for order, variable in enumerate(all_columns, start=1):
    in_2023 = variable in columns_by_year[2023]
    in_2024 = variable in columns_by_year[2024]
    in_2025 = variable in columns_by_year[2025]

    available_year_count = sum([in_2023, in_2024, in_2025])

    feature_rows.append({
        "order": order,
        "variable": variable,
        "final_label": choose_label(variable),
        "label_2023": labels_by_year[2023].get(variable, ""),
        "label_2024": labels_by_year[2024].get(variable, ""),
        "label_2025": labels_by_year[2025].get(variable, ""),
        "label_consistency": label_consistency(variable),
        "in_2023": in_2023,
        "in_2024": in_2024,
        "in_2025": in_2025,
        "available_year_count": available_year_count,
        "is_common_3years": available_year_count == 3,
        "domain_guess": infer_domain(variable),
        "initial_decision": infer_initial_decision(variable),
    })

feature_catalog_df = pd.DataFrame(feature_rows)


# ============================================================
# 6. 값 라벨표 생성
# ============================================================
value_label_rows = []

for variable in all_columns:
    for year in sorted(value_labels_by_year):
        value_label_dict = value_labels_by_year[year].get(variable)

        if not value_label_dict:
            continue

        for value, label in value_label_dict.items():
            value_label_rows.append({
                "year": year,
                "variable": variable,
                "variable_label": labels_by_year[year].get(variable, ""),
                "value": value,
                "value_label": label,
            })

value_label_df = pd.DataFrame(value_label_rows)


# ============================================================
# 7. 전체 원자료 읽기 + current_ecig_use 생성 + 합집합 병합 준비
# ============================================================
combined_dfs = []
year_summary_rows = []

for year, path in SAV_FILES.items():
    print(f"\n{year}년 원자료 읽는 중: {path.name}")

    df, meta = pyreadstat.read_sav(str(path))

    print(f"{year}년 원본 크기:", df.shape)

    # 파일 기준 연도 명시
    if "YEAR" in df.columns:
        df = df.drop(columns=["YEAR"])

    df.insert(0, "YEAR", year)

    if TARGET_SOURCE not in df.columns:
        raise ValueError(f"{year}년 데이터에 {TARGET_SOURCE} 변수가 없습니다.")

    df[TARGET_NAME] = df[TARGET_SOURCE].apply(make_current_ecig_use)

    invalid_target_count = int(df[TARGET_NAME].isna().sum())

    # 전체 원자료 병합본에서는 TC_EC_MN을 보존한다.
    # 다만 생성 타깃은 정수형으로 변환 가능한 행만 남긴다.
    before_drop = len(df)
    df = df.dropna(subset=[TARGET_NAME])
    after_drop = len(df)

    df[TARGET_NAME] = df[TARGET_NAME].astype(int)

    y_counts = df[TARGET_NAME].value_counts().sort_index()
    y_ratios = df[TARGET_NAME].value_counts(normalize=True).sort_index()

    year_summary_rows.append({
        "year": year,
        "original_rows": before_drop,
        "processed_rows": after_drop,
        "dropped_rows": before_drop - after_drop,
        "original_column_count": len(meta.column_names),
        "invalid_target_count": invalid_target_count,
        "current_non_user_y0_count": int(y_counts.get(0, 0)),
        "current_user_y1_count": int(y_counts.get(1, 0)),
        "current_non_user_y0_ratio": float(y_ratios.get(0, 0)),
        "current_user_y1_ratio": float(y_ratios.get(1, 0)),
    })

    combined_dfs.append(df)


# ============================================================
# 8. 합집합 기준 전체 병합
# ============================================================
full_combined_df = pd.concat(
    combined_dfs,
    axis=0,
    join="outer",
    ignore_index=True,
    sort=False,
)

# YEAR를 첫 번째, current_ecig_use를 마지막으로 배치
cols = list(full_combined_df.columns)
cols_without_year_target = [
    col for col in cols
    if col not in ["YEAR", TARGET_NAME]
]

full_combined_df = full_combined_df[
    ["YEAR"] + cols_without_year_target + [TARGET_NAME]
]

print("\n전체 병합 데이터 크기:", full_combined_df.shape)

print("\n연도별 행 수:")
print(full_combined_df["YEAR"].value_counts().sort_index())

print("\n타깃 분포:")
print(full_combined_df[TARGET_NAME].value_counts().sort_index())


# ============================================================
# 9. 변수별 결측/9999/8888 요약
# ============================================================
missing_summary_rows = []

for variable in full_combined_df.columns:
    series = full_combined_df[variable]

    total_count = len(series)
    na_count = int(series.isna().sum())

    count_9999 = 0
    count_8888 = 0

    if pd.api.types.is_numeric_dtype(series):
        count_9999 = int((series == 9999).sum())
        count_8888 = int((series == 8888).sum())

    missing_summary_rows.append({
        "variable": variable,
        "final_label": choose_label(variable) if variable != TARGET_NAME else "현재 액상형 전자담배 사용 여부",
        "na_count": na_count,
        "na_ratio": na_count / total_count,
        "count_9999": count_9999,
        "ratio_9999": count_9999 / total_count,
        "count_8888": count_8888,
        "ratio_8888": count_8888 / total_count,
        "dtype": str(series.dtype),
        "n_unique_dropna": int(series.nunique(dropna=True)),
    })

missing_summary_df = pd.DataFrame(missing_summary_rows)


# ============================================================
# 10. 요약표 생성
# ============================================================
year_summary_df = pd.DataFrame(year_summary_rows)

target_distribution_df = (
    full_combined_df[TARGET_NAME]
    .value_counts()
    .sort_index()
    .rename_axis(TARGET_NAME)
    .reset_index(name="count")
)

target_distribution_df["ratio"] = (
    target_distribution_df["count"] / len(full_combined_df)
)

domain_summary_df = (
    feature_catalog_df
    .groupby(["domain_guess", "initial_decision"], dropna=False)
    .size()
    .reset_index(name="variable_count")
    .sort_values(["domain_guess", "initial_decision"])
)

readme_df = pd.DataFrame([
    {
        "item": "파일 목적",
        "content": "2023~2025년 청소년건강행태조사 원자료의 전체 변수 합집합 병합본 및 전체 feature 목록 확인",
    },
    {
        "item": "병합 방식",
        "content": "연도별 데이터 세로 결합, 컬럼은 합집합(join='outer') 기준",
    },
    {
        "item": "원본 파일",
        "content": "kyrbs2023.sav, kyrbs2024.sav, kyrbs2025.sav",
    },
    {
        "item": "생성 타깃",
        "content": "current_ecig_use",
    },
    {
        "item": "타깃 기준",
        "content": "TC_EC_MN=2~7이면 1, TC_EC_MN=1 또는 9999이면 0",
    },
    {
        "item": "주의",
        "content": "이 파일은 전체 탐색용이며, 모델링용 feature는 별도로 선별해야 함",
    },
])


# ============================================================
# 11. 파일 저장
# ============================================================
full_combined_df.to_csv(
    FULL_CSV_FILE,
    index=False,
    encoding="utf-8-sig",
)

print("\n전체 병합 CSV 저장 완료:", FULL_CSV_FILE)

# 전체 데이터 Excel은 크기가 커서 시간이 조금 걸릴 수 있다.
with pd.ExcelWriter(FULL_XLSX_FILE, engine="openpyxl") as writer:
    full_combined_df.to_excel(
        writer,
        sheet_name="full_combined_data",
        index=False,
        startrow=1,
    )

print("전체 병합 Excel 저장 완료:", FULL_XLSX_FILE)

with pd.ExcelWriter(FEATURE_CATALOG_FILE, engine="openpyxl") as writer:
    readme_df.to_excel(writer, sheet_name="README", index=False)
    feature_catalog_df.to_excel(writer, sheet_name="feature_catalog", index=False)
    value_label_df.to_excel(writer, sheet_name="value_labels", index=False)
    missing_summary_df.to_excel(writer, sheet_name="missing_summary", index=False)
    year_summary_df.to_excel(writer, sheet_name="year_summary", index=False)
    target_distribution_df.to_excel(writer, sheet_name="target_distribution", index=False)
    domain_summary_df.to_excel(writer, sheet_name="domain_summary", index=False)

print("전체 feature catalog 저장 완료:", FEATURE_CATALOG_FILE)


# ============================================================
# 12. 전체 Excel 상단에 한글 라벨 행 추가
# ============================================================
wb = load_workbook(FULL_XLSX_FILE)
ws = wb["full_combined_data"]

header_fill = PatternFill("solid", fgColor="1F4E78")
variable_fill = PatternFill("solid", fgColor="D9EAF7")
white_font = Font(color="FFFFFF", bold=True)
bold_font = Font(bold=True)

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

for col_idx, variable_name in enumerate(full_combined_df.columns, start=1):
    if variable_name == TARGET_NAME:
        korean_label = "현재 액상형 전자담배 사용 여부"
    else:
        korean_label = choose_label(variable_name)

    label_cell = ws.cell(row=1, column=col_idx)
    label_cell.value = korean_label
    label_cell.fill = header_fill
    label_cell.font = white_font
    label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_cell.border = thin_border

    variable_cell = ws.cell(row=2, column=col_idx)
    variable_cell.fill = variable_fill
    variable_cell.font = bold_font
    variable_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    variable_cell.border = thin_border

    max_len = max(len(str(variable_name)), len(str(korean_label)))
    col_letter = ws.cell(row=2, column=col_idx).column_letter
    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 35)

ws.freeze_panes = "A3"
ws.auto_filter.ref = ws.dimensions
ws.row_dimensions[1].height = 42
ws.row_dimensions[2].height = 28

wb.save(FULL_XLSX_FILE)

print("전체 병합 Excel 라벨 행 추가 완료:", FULL_XLSX_FILE)


# ============================================================
# 13. 최종 출력
# ============================================================
print("\n===== 생성 완료 =====")
print("1. 전체 원자료 병합 CSV:")
print("   ", FULL_CSV_FILE)

print("2. 전체 원자료 병합 Excel:")
print("   ", FULL_XLSX_FILE)

print("3. 전체 feature 확인표:")
print("   ", FEATURE_CATALOG_FILE)

print("\n전체 병합 데이터 크기:", full_combined_df.shape)
print("전체 feature 합집합 개수:", len(all_columns))
print("3개년 공통 feature 개수:", len(common_columns))

print("\n연도별 요약:")
print(year_summary_df.to_string(index=False))

print("\n타깃 분포:")
print(target_distribution_df.to_string(index=False))
